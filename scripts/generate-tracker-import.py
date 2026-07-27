#!/usr/bin/env python3
"""Regenerate data/tracker-import.json from the cleaned Tracker Excel sheet."""

from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "Bioinfo Service Reports Reference Number_CLEANED.xlsx"
OUT = ROOT / "data" / "tracker-import.json"

CLASSIFICATION_ALIASES = {
    "wgs analyses": "Whole Genome Assembly",
    "wgs analysis": "Whole Genome Assembly",
    "whole genome assembly": "Whole Genome Assembly",
    "16s metabarcoding": "16s Metabarcoding",
    "amplicon": "Amplicon",
    "edna analysis": "eDNA Analysis",
    "phylogenetics": "Phylogenetics",
    "transcriptomics": "Transcriptomics",
    "capseq": "CapSeq",
    "mtdna": "mtDNA",
    "cpdna": "cpDNA",
    "shotgun metagenomics": "Shotgun Metagenomics",
    "population genetics": "Population Genetics",
    "others": "Others",
}
OPTIONS = set(CLASSIFICATION_ALIASES.values()) | {
    "Amplicon",
    "Whole Genome Assembly",
    "16s Metabarcoding",
    "eDNA Analysis",
    "Phylogenetics",
    "Transcriptomics",
    "CapSeq",
    "mtDNA",
    "cpDNA",
    "Shotgun Metagenomics",
    "Population Genetics",
    "Others",
}


def cell_str(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    s = str(v).strip()
    return s if s else None


def normalize_date(v):
    """Return YYYY-MM-DD or None. Reject year-only / invalid Excel values."""
    raw = cell_str(v)
    if not raw:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    # Already ISO date
    if len(raw) >= 10 and raw[4] == "-" and raw[7] == "-":
        return raw[:10]
    # Year-only or other junk from Excel (e.g. "2025")
    return None


def format_sr(prefix, suffix):
    p = cell_str(prefix)
    if not p:
        return None
    if suffix is None or (isinstance(suffix, str) and not str(suffix).strip()):
        return p
    try:
        n = int(float(suffix))
        return f"{p}-{n:03d}"
    except Exception:
        return f"{p}-{str(suffix).strip()}"


def normalize_class(raw):
    if not raw:
        return None
    t = raw.strip()
    low = t.lower()
    if low in CLASSIFICATION_ALIASES:
        return CLASSIFICATION_ALIASES[low]
    for o in OPTIONS:
        if o.lower() == low:
            return o
    return t


def map_status(label):
    if not label:
        return None
    t = label.strip().lower()
    if t == "completed":
        return "completed"
    if t in ("on-going", "ongoing", "on going"):
        return "ongoing"
    if t == "submitted":
        return "submitted"
    if t in ("for approval", "for_approval"):
        return "for_approval"
    if "on hold" in t or t == "on_hold":
        return "on_hold"
    return None


def derive_legacy(comp, sub, ana):
    return map_status(comp) or map_status(sub) or map_status(ana) or "for_approval"


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Tracker"]
    all_rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
    last_nonblank = 1
    for i, row in enumerate(all_rows, start=2):
        if any(
            v is not None and not (isinstance(v, str) and not str(v).strip())
            for v in row
        ):
            last_nonblank = i

    rows_out = []
    for i, row in enumerate(all_rows, start=2):
        if i > last_nonblank:
            break
        application = cell_str(row[3])
        classification = cell_str(row[4])
        pipeline = normalize_class(classification)
        if application:
            pipeline = "Others"
        sr_date = normalize_date(row[2])
        status_analysis = cell_str(row[11])
        status_completion = cell_str(row[12])
        status_submission = cell_str(row[13])
        sequences_link = cell_str(row[15])
        legacy = derive_legacy(status_completion, status_submission, status_analysis)
        completed_at = f"{sr_date}T00:00:00.000Z" if legacy == "completed" and sr_date else None
        rows_out.append(
            {
                "excel_row": i,
                "service_report_number": format_sr(row[0], row[1]),
                "service_report_date": sr_date,
                "application": application,
                "pipeline": pipeline,
                "client_name": cell_str(row[5]),
                "client_type": cell_str(row[6]),
                "external_client_id": cell_str(row[7]),
                "external_project_id": cell_str(row[8]),
                "sample_type": cell_str(row[9]),
                "run_id": cell_str(row[10]),
                "status_of_analysis": status_analysis,
                "status_of_completion": status_completion,
                "status_of_submission": status_submission,
                "service_report_link": cell_str(row[14]),
                "client_sequences_link": sequences_link,
                "notes": cell_str(row[16]),
                "status": legacy,
                "started_at": f"{sr_date}T00:00:00.000Z" if sr_date else None,
                "completed_at": completed_at,
                "project_id": None,
                "assignee_id": None,
                "pipeline_version": None,
                "output_link": sequences_link,
            }
        )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(
        json.dumps(
            {"source": "Tracker sheet", "count": len(rows_out), "records": rows_out},
            indent=2,
        )
        + "\n"
    )
    print(f"Wrote {len(rows_out)} records → {OUT}")


if __name__ == "__main__":
    main()